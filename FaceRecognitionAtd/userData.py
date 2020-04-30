# -*- coding: utf-8 -*-
"""
Created on Mon Mar 23 17:51:15 2020

@author: Chaitanya
"""

import openpyxl
import os

filename = 'userdata.xlsx'

if os.path.isfile(filename):
    wb = openpyxl.Workbook(filename)
    if 'Data' in wb.get_sheet_names():
        pass
    else:
        wb.create_sheet(index=0, title='Data')
else:
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Data')
    wb.save(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Data')


def add_user(name, sccode, branch):
    ws = wb.active
    first_column = ws['A']
    second_column = ws['B']
    third_column = ws['C']
    col_len1 = str(len(first_column) + 1)
    col_len2 = str(len(second_column) + 1)
    col_len3 = str(len(third_column) + 1)
    sheet['A' + col_len1] = name
    sheet['B' + col_len2] = sccode
    sheet['C' + col_len3] = branch
    wb.save(filename)


if (sheet['A1'].value == 'Name') and (sheet['B1'].value == 'SC Code') and (sheet['C1'].value == 'Branch'):
    pass
else:
    sheet['A1'] = 'Name'
    sheet['B1'] = 'SC Code'
    sheet['C1'] = 'Branch'
